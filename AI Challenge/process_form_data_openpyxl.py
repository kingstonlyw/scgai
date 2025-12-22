import argparse
import json
from pathlib import Path
from typing import Dict, List, Any

from openpyxl import load_workbook
from datetime import datetime, date



def ensure_parent_dir(path: Path) -> None:
    if path.parent and not path.parent.exists():
        path.parent.mkdir(parents=True, exist_ok=True)


def load_rename_mapping(path: str | None) -> Dict[str, str]:
    if not path:
        return {}
    with open(path, "r", encoding="utf-8") as f:
        mapping = json.load(f)
    if not isinstance(mapping, dict):
        raise ValueError("--rename must be a JSON object mapping source->target column names")
    return mapping


def trim_value(v: Any) -> Any:
    if isinstance(v, str):
        return v.strip()
    return v


def is_empty_row(values: List[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def format_date_value(v: Any, fmt: str) -> Any:
    if isinstance(v, datetime):
        return v.strftime(fmt)
    if isinstance(v, date):
        # date without time
        return datetime(v.year, v.month, v.day).strftime(fmt)
    # If it’s a string, we avoid heavy parsing deps; leave as-is
    return v


# def main():
#     parser = argparse.ArgumentParser(description="Convert Excel form data to structured JSON (openpyxl only)")
#     parser.add_argument("--excel", required=True, help="Path to the input .xlsx file")
#     parser.add_argument("--sheet", default=0, help="Sheet name or index (default: 0)")
#     parser.add_argument(
#         "--header-row",
#         type=int,
#         default=1,
#         help="1-based header row number containing column names (default: 1)",
#     )
#     parser.add_argument("--rename", help="Path to JSON file mapping source column names to output names")
#     parser.add_argument("--date-cols", nargs="*", default=[], help="Column names to format as dates")
#     parser.add_argument("--date-format", default="%Y-%m-%dT%H:%M:%S", help="strftime format for dates")
#     parser.add_argument("--filter-eq", nargs="*", default=[], help="Equality filters like Status=Active")
#     parser.add_argument("--dropna", action="store_true", help="Drop rows that are entirely empty")
#     parser.add_argument("--output", default="output/data.json", help="Output JSON path")
#     parser.add_argument("--pretty", action="store_true", help="Pretty-print JSON with indentation")

#     args = parser.parse_args()

#     excel_path = Path(args.excel)
#     if not excel_path.exists():
#         raise SystemExit(f"Excel file not found: {excel_path}")

#     wb = load_workbook(filename=str(excel_path), data_only=True, read_only=True)

#     # Resolve sheet
#     if isinstance(args.sheet, str):
#         try:
#             sheet_index = int(args.sheet)
#             ws = wb.worksheets[sheet_index]
#         except ValueError:
#             if args.sheet not in wb.sheetnames:
#                 raise SystemExit(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
#             ws = wb[args.sheet]
#     else:
#         ws = wb.worksheets[int(args.sheet)]

#     header_row = max(args.header_row, 1)
#     header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
#     headers = [str(h).strip() if h is not None else "" for h in header_cells]

#     rename_map = load_rename_mapping(args.rename)
#     date_cols_set = set(args.date_cols or [])

#     out_path = Path(args.output)
#     ensure_parent_dir(out_path)

#     # Build records
#     records = []
#     for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
#         values = list(row)
#         if args.dropna and is_empty_row(values):
#             continue
#         obj: Dict[str, Any] = {}
#         for idx, raw_val in enumerate(values):
#             if idx >= len(headers):
#                 continue
#             col_name = headers[idx]
#             if not col_name:
#                 continue
#             out_key = rename_map.get(col_name, col_name)
#             val = trim_value(raw_val)
#             if col_name in date_cols_set:
#                 val = format_date_value(val, args.date_format)
#             obj[out_key] = val

#         records.append(obj)

#     # Filters
#     filters: Dict[str, str] = {}
#     for f in args.filter_eq or []:
#         if "=" not in f:
#             raise SystemExit(f"Invalid --filter-eq '{f}'. Use Column=Value")
#         k, v = f.split("=", 1)
#         filters[k.strip()] = v.strip()

#     if filters:
#         def match(rec: Dict[str, Any]) -> bool:
#             for k, v in filters.items():
#                 # Support both pre-rename and post-rename keys
#                 actual_val = rec.get(k)
#                 if actual_val is None:
#                     # Try renamed key
#                     rk = rename_map.get(k, None)
#                     if rk is not None:
#                         actual_val = rec.get(rk)
#                 if str(actual_val) != str(v):
#                     return False
#             return True

#         records = [r for r in records if match(r)]

#         # Load existing records if submissions.json exists
#         existing_records = []
#         existing_set = set()

#         if out_path.exists():
#             with open(out_path, "r", encoding="utf-8") as f:
#                 existing_records = json.load(f)

#             existing_set = {
#                 json.dumps(r, sort_keys=True, ensure_ascii=False)
#                 for r in existing_records
#             }

#         # Keep only new records
#         new_records = []
#         for r in records:
#             key = json.dumps(r, sort_keys=True, ensure_ascii=False)
#             if key not in existing_set:
#                 new_records.append(r)
#                 existing_set.add(key)

#     # Append new records to existing ones
#     records = existing_records + new_records

#     indent = 2 if args.pretty else None
#     with open(out_path, "w", encoding="utf-8") as f:
#         json.dump(records, f, ensure_ascii=False, indent=indent)

#     print(f"Wrote {len(records)} records to {out_path}")

def main():
    parser = argparse.ArgumentParser(description="Convert Excel form data to structured JSON (openpyxl only)")
    parser.add_argument("--excel", required=True, help="Path to the input .xlsx file")
    parser.add_argument("--sheet", default=0, help="Sheet name or index (default: 0)")
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="1-based header row number containing column names (default: 1)",
    )
    parser.add_argument("--rename", help="Path to JSON file mapping source column names to output names")
    parser.add_argument("--date-cols", nargs="*", default=[], help="Column names to format as dates")
    parser.add_argument("--date-format", default="%Y-%m-%dT%H:%M:%S", help="strftime format for dates")
    parser.add_argument("--filter-eq", nargs="*", default=[], help="Equality filters like Status=Active")
    parser.add_argument("--dropna", action="store_true", help="Drop rows that are entirely empty")
    parser.add_argument("--output", default="output/data.json", help="Output JSON path")
    parser.add_argument("--pretty", action="store_true", help="Pretty-print JSON with indentation")

    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    wb = load_workbook(filename=str(excel_path), data_only=True, read_only=True)

    # Resolve sheet
    if isinstance(args.sheet, str):
        try:
            sheet_index = int(args.sheet)
            ws = wb.worksheets[sheet_index]
        except ValueError:
            if args.sheet not in wb.sheetnames:
                raise SystemExit(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
            ws = wb[args.sheet]
    else:
        ws = wb.worksheets[int(args.sheet)]

    header_row = max(args.header_row, 1)
    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_cells]

    rename_map = load_rename_mapping(args.rename)
    date_cols_set = set(args.date_cols or [])

    out_path = Path(args.output)
    ensure_parent_dir(out_path)

    # Build records
    records: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = list(row)
        if args.dropna and is_empty_row(values):
            continue

        obj: Dict[str, Any] = {}
        for idx, raw_val in enumerate(values):
            if idx >= len(headers):
                continue
            col_name = headers[idx]
            if not col_name:
                continue

            out_key = rename_map.get(col_name, col_name)
            val = trim_value(raw_val)
            if col_name in date_cols_set:
                val = format_date_value(val, args.date_format)
            obj[out_key] = val

        records.append(obj)

    # Filters
    filters: Dict[str, str] = {}
    for f in args.filter_eq or []:
        if "=" not in f:
            raise SystemExit(f"Invalid --filter-eq '{f}'. Use Column=Value")
        k, v = f.split("=", 1)
        filters[k.strip()] = v.strip()

    if filters:
        def match(rec: Dict[str, Any]) -> bool:
            for k, v in filters.items():
                # Support both pre-rename and post-rename keys
                actual_val = rec.get(k)
                if actual_val is None:
                    rk = rename_map.get(k, None)
                    if rk is not None:
                        actual_val = rec.get(rk)
                if str(actual_val) != str(v):
                    return False
            return True

        records = [r for r in records if match(r)]

    # Deduplicate/append based on ID + Start time, using renamed keys if rename_map was applied
    id_key = rename_map.get("ID", "ID")
    start_key = rename_map.get("Start time", "Start time")

    existing_records: List[Dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    if out_path.exists():
        try:
            with open(out_path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, list):
                existing_records = [r for r in loaded if isinstance(r, dict)]
        except Exception:
            existing_records = []

        for r in existing_records:
            rid = r.get(id_key)
            st = r.get(start_key)
            if rid is None or st is None:
                continue
            seen.add((str(rid), str(st)))

    new_records: List[Dict[str, Any]] = []
    for r in records:
        rid = r.get(id_key)
        st = r.get(start_key)

        # simplest behavior: if either is missing, keep it (can't dedupe reliably)
        if rid is None or st is None:
            new_records.append(r)
            continue

        key = (str(rid), str(st))
        if key not in seen:
            new_records.append(r)
            seen.add(key)

    records = existing_records + new_records

    indent = 2 if args.pretty else None
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=indent)

    print(f"Wrote {len(records)} records to {out_path}")

if __name__ == "__main__":
    main()
